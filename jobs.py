import logging
import traceback
import os
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from ai import *
from jobs_scraper import *



excel_file = "jobs.xlsx"
required_columns = ["applied", "AI_recommendation", "AI_explanation", "company", "title", "link", "description",
                    "scrape_date",
                    "posted_date"]
pd.set_option('display.max_rows', None)  # Show all rows
pd.set_option('display.max_columns', None)  # Show all columns
pd.set_option('display.width', None)
logging.basicConfig(level=logging.INFO)


def load_env_file(file_path):
    with open(file_path) as f:
        for line in f:
            # Ignore comments and empty lines
            if line.strip() and not line.startswith('#') and not line.startswith(';'):
                key, value = line.strip().split('=', 1)
                os.environ[key] = value


def load_df():
    if os.path.exists(excel_file):
        df = pd.read_excel(excel_file, engine='openpyxl')
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Missing required column: {col}")
    else:
        df = pd.DataFrame(columns=required_columns)
    # df.to_csv("after_ai_temp.csv", index=False)
    return df


def split_string(input_string):
    starting_words = ["yes", "no", "maybe+", "maybe"]
    for word in starting_words:
        if input_string.startswith(word):
            return word, input_string[len(word):].strip()
    logging.error(f"Invalid answer of AI: doesnt start with {starting_words}")
    return "", input_string


def beautify_excel():
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.auto_filter.ref = ws.dimensions
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):

        for cell in row:
            if cell.column_letter in ["A", "B"]:
                # dv.add(cell)
                if cell.value == "no":
                    cell.fill = red_fill
                elif cell.value == "yes" or cell.value == "maybe+":
                    cell.fill = green_fill
                elif cell.value not in ["yes", "no"]:
                    cell.fill = yellow_fill
            if cell.column_letter == "F":
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    wb.save(excel_file)


def general_scrape_and_ai(unique_urls, assistant):
    """
    scrape and filter.
    The script runs in a loop until all data is scraped (to handle cases where the process may be blocked midway,
    such as by LinkedIn). By default, there is a break statement in the loop, which stops the process after a single
    attempt. If you want the process to continue scraping until all data is collected, you should remove the break
    statement. Keep in mind, if the program fails midway, there is a risk of data loss.
    :param unique_urls: set of all urls already scraped and filtered
    :param assistant: AI assistant
    :return: df with all new scraped job including the AI filter
    """
    offset = 0
    new_data = [0]
    df = pd.DataFrame(columns=required_columns)

    while len(new_data) > 0:
        try:
            logging.info(f"Offset of scraping is: {offset}")
            new_data = scrape_all_jobs(os.getenv('sites'), os.getenv('search_term'), os.getenv('location'),
                                       os.getenv('hours_old'),
                                       os.getenv('results_wanted'), offset)
            logging.info(f"{len(new_data)} jobs scraped ")
            # new_data.to_csv(f"temp{offset}.csv", index=False)
            # new_data = pd.read_csv("temp.csv")
        except Exception as e:
            logging.error("An error occurred while scraping: %s", e)
            logging.error("Stack trace: %s", traceback.format_exc())
            new_data = pd.DataFrame(columns=required_columns)

        for index, row in new_data.iterrows():
            try:
                if row['job_url'] in unique_urls:
                    continue
                msg = f"title: {row['title']}. description: {row['description']}"
                ai_response = assistant.submit_message(msg)
                # print(ai_recommend)
                recommendation, explanation = split_string(ai_response)
                new_row = {
                    "applied": "",
                    "AI_recommendation": recommendation,
                    "AI_explanation": explanation,
                    "company": row['company'],
                    "title": row['title'],
                    "link": row['job_url'],
                    "description": row['description'],
                    "posted_date": row['date_posted'],
                    "scrape_date": date.today(),

                }
                df.loc[len(df)] = new_row
                # new_row_df = pd.DataFrame([new_row])
                # new_row_df.to_csv(f"after_ai_temp.csv", mode='a', index=False, header=False)
                unique_urls.add(new_row['link'])
            except Exception as e:
                logging.error("An error occurred while sending to gpt: %s", e)
                logging.error("Stack trace: %s", traceback.format_exc())
        offset += len(new_data)
        break

    return df


def main():
    load_env_file('.env')
    data = load_df()
    unique_urls = set(data["link"])
    assistant_name = "MyJobsMatcher"
    with open('instructions.txt', 'r', encoding='utf-8') as file:
        instructions = file.read()
    assistant = OpenAIAssistant(os.getenv('api_key'), assistant_name, instructions, os.getenv('model'),
                                os.getenv('assistant_id'))
    # for using the same assistant for the next round
    logging.info(f"Assistant created successfully. assistant ID: {assistant.assistant_id}")
    new_df = general_scrape_and_ai(unique_urls, assistant)
    df = pd.concat([data, new_df])
    df.to_excel(excel_file, index=False, engine='openpyxl')
    beautify_excel()


if __name__ == "__main__":
    main()
